from sqlalchemy import create_engine
import pandas as pd
from openpyxl import load_workbook
from os import walk
from Class_Data.Command import Filter
import bu_module as bu_md

DEBUG_ON = 1
DEBUG_OFF = 0

debug = DEBUG_OFF
D_read_sheets = DEBUG_OFF
D_create_new_excel = DEBUG_OFF
D_load_spr_info_excel = DEBUG_ON
D_make_df_spr_info = DEBUG_OFF
D_load_spr_in_folds = DEBUG_ON
D_make_df_from_arr = DEBUG_OFF
D_save_filtered_spr_info_excel = DEBUG_ON
D_save_sum_spr_excel = DEBUG_OFF
D_load_spr_sum_excel = DEBUG_OFF
D_save_filtered_spr_sum_excel = DEBUG_OFF

D_load_llp_sum_spr_excel = DEBUG_OFF


D_load_time_excel = DEBUG_OFF
D_load_cur_plan_excel = DEBUG_OFF
D_make_df_plan_sum = DEBUG_OFF
D_make_df_plan_jobs = DEBUG_OFF

D_update_df_from_arr = DEBUG_OFF
D_save_time_excel = DEBUG_OFF
D_connect_sql_server = DEBUG_OFF
D_save_time_excel_to_server = DEBUG_OFF
D_load_time_excel_from_server = DEBUG_OFF


# FUNCTION CODE : save_filtered_spr_info_excel
# comment : 테이블에 떠있는 정보를 DATA FRAME 으로 내보낸다.
###############################################################################################


def write_qtable_to_df(table):
    col_count = table.columnCount()
    row_count = table.rowCount()
    headers = [str(table.horizontalHeaderItem(i).text()) for i in range(col_count)]

    # df indexing is slow, so use lists
    df_list = []
    for row in range(row_count):
        df_list2 = []
        for col in range(col_count):
            table_item = table.item(row, col)
            df_list2.append("" if table_item is None else str(table_item.text()))
        df_list.append(df_list2)

    df = pd.DataFrame(df_list, columns=headers)

    return df


# FUNCTION CODE : save_filtered_spr_sum_excel
# comment : 테이블에 떠있는 SPR SUM 데이터를 EXCEL 파일로 저장한다.
###############################################################################################
def save_filtered_spr_sum_excel(self):
    drv_path = "D:"
    drv_path = r"Y:\TEST"
    src = (
        drv_path
        + r"\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM\00. SPR FILTERED DATA/"
    )

    if self.ed2_1.text():
        pathner = "." + self.ed2_1.text() + "."
    else:
        pathner = ""

    if self.ed2_2.text():
        oem = self.ed2_2.text() + "."
    else:
        oem = ""

    if self.ed2_3.text():
        lges = self.ed2_3.text() + "."
    else:
        lges = ""

    if self.ed2_4.text():
        ps = self.ed2_4.text() + "."
    else:
        ps = ""

    file = f"SUM_SPR_FILTERED{pathner}{oem}{lges}{ps}"

    if not (pathner and oem and lges and ps):
        src_path = src + file + ".xlsx"
    else:
        src_path = src + file[0:-1] + ".xlsx"

    try:
        df = write_qtable_to_df(self.tbw1)
        df["LLP"] = df["LLP"].apply(lambda x: float(x))
        df["Factor"] = df["Factor"].apply(lambda x: float(x))
        df["Price"] = df["Price"].apply(lambda x: float(x))
        df["Qty"] = df["Qty"].apply(lambda x: float(x))
        df["Total"] = df["Total"].apply(lambda x: float(x))
        df["Dis"] = df["Dis"].apply(lambda x: float(x))
        df["금액"] = df["금액"].apply(lambda x: float(x))

        if debug or D_save_filtered_spr_sum_excel:
            SUCCESS = f"save_filtered_spr_sum_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")

        df.to_excel(src_path, sheet_name="SPR", index=False)

    except:
        if debug or D_save_filtered_spr_sum_excel:
            ERROR = f"save_filtered_spr_sum_excel : ERROR 타임시트 엑셀에 저장실패"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : save_sum_spr_excel
# comment : 테이블에 떠있는 SPR INFO 데이터를 EXCEL 파일로 저장한다.
###############################################################################################
def save_sum_spr_excel(self):
    src = r"D:\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM/"
    src = r"Y:\TEST\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM/"

    file = f"SUM_SPR_DATA"
    src_path = src + file + ".xlsx"

    try:
        df = write_qtable_to_df(self.tbw1)
        df["LLP"] = df["LLP"].apply(lambda x: float(x))
        df["Factor"] = df["Factor"].apply(lambda x: float(x))
        df["Price"] = df["Price"].apply(lambda x: float(x))
        df["Qty"] = df["Qty"].apply(lambda x: float(x))
        df["Total"] = df["Total"].apply(lambda x: float(x))
        df["Dis"] = df["Dis"].apply(lambda x: float(x))

        if debug or D_save_sum_spr_excel:
            SUCCESS = f"save_sum_spr_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")

        df.to_excel(src_path, sheet_name="SPR", index=False)

    except:
        if debug or D_save_sum_spr_excel:
            ERROR = f"save_sum_spr_excel : ERROR 타임시트 엑셀에 저장실패"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : save_filtered_spr_info_excel
# comment : 테이블에 떠있는 SPR INFO 데이터를 EXCEL 파일로 저장한다.
###############################################################################################
def save_filtered_spr_info_excel(self):
    # src = r"D:\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\001. SPR INFO\00. SPR FILTER INFO/"
    src = r"Y:\TEST\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\001. SPR INFO\00. SPR FILTER INFO/"
    if self.ed2_1.text():
        pathner = "." + self.ed2_1.text() + "."
    else:
        pathner = ""

    if self.ed2_2.text():
        oem = self.ed2_2.text() + "."
    else:
        oem = ""

    if self.ed2_3.text():
        lges = self.ed2_3.text() + "."
    else:
        lges = ""

    if self.ed2_4.text():
        ps = self.ed2_4.text() + "."
    else:
        ps = ""

    file = f"SPR_PRG_INFO{pathner}{oem}{lges}{ps}"

    if not (pathner and oem and lges and ps):
        src_path = src + file + ".xlsx"
    else:
        src_path = src + file[0:-1] + ".xlsx"

    try:
        df = write_qtable_to_df(self.tbw1)
        df["금액"] = df["금액"].apply(lambda x: int(x))

        if debug or D_save_filtered_spr_info_excel:
            SUCCESS = f"save_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")

        df.to_excel(src_path, sheet_name="INFO", index=False)

    except:
        if debug or D_save_filtered_spr_info_excel:
            ERROR = f"save_time_excel : ERROR 타임시트 엑셀에 저장실패"
            print(f"CODE : {ERROR} ")
            print(f"CODE : {df} ")


# FUNCTION CODE = read_sheets()
# path         : 파일이 있는 경로를 지정한다.
# excel        : path +엑셀파일.xlsx
# return        : sheet 명을 배열 형태로 돌려준다.
# comment      : 독립적으로 동작하고, 경로와, 엑셀파일을 넣으면 해당하는 모든 시트를 읽는다.
###############################################################################################
def read_sheets(path, excel):
    i = -1
    wb = load_workbook(path + excel)
    shts = []
    try:
        for item in wb.sheetnames:
            i = i + 1
            shts.append(item)
        if debug or D_read_sheets:
            SUCCESS = f"read_sheets : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {shts} ")
        return shts
    except:
        if debug or D_read_sheets:
            ERROR = f"read_sheets : ERROR 파일확장자가 지정되지않음.EX) .xlsx"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : create_new_excel
# path = 이 안에 들어 있는 모든 엑셀 파일의 타겟 sheet를 삭제한다.
# file        : path +엑셀파일.xlsx
# tar_sht_number : 시트번호, 첫번째 0, 두번째 1 ... n번째 n
# comment      : 독립적으로 동작하고, 엑셀파일과 시트번호을 넣으면, 해당하는 시트가 삭제된다.
###############################################################################################
def create_new_excel():
    path = r"D:\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM/"
    path = r"Y:\TEST\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM/"
    file = "SPR_DATA.xlsx"
    df = pd.DataFrame()
    sht = ["MASTER", "DATA", "RAW"]
    try:
        with pd.ExcelWriter(path + file) as writer:
            df.to_excel(writer, sheet_name=sht[0], index=False)
            df.to_excel(writer, sheet_name=sht[1], index=False)
            df.to_excel(writer, sheet_name=sht[2], index=False)
        if debug or D_create_new_excel:
            SUCCESS = f"create_new_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            read_sheets(path, file)

    except:
        if debug or D_create_new_excel:
            ERROR = f"create_new_excel : ERROR 파일확장자가 지정되지않음.EX) .xlsx"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_spr_sum_excel
#
# comment : 필터적용안된 SPR 데이터를 불러옴
###############################################################################################
def load_spr_sum_excel(self):
    # Y:\TEST\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM
    drv_path = "D:"
    drv_path = r"Y:\TEST"
    path_spr_info = (
        drv_path
        + r"\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\001. SPR INFO\00. SPR FILTER INFO/"
    )
    file_spr_info = "SPR_PRG_INFO.xlsx"
    src_spr_info = path_spr_info + file_spr_info

    path_spr_sum = drv_path + r"\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM/"
    file_spr_sum = "SUM_SPR_DATA.xlsx"
    src_spr_sum = path_spr_sum + file_spr_sum

    try:
        df_spr_info = pd.read_excel(src_spr_info, sheet_name=0)
        df_spr_sum = pd.read_excel(src_spr_sum, sheet_name=0)
        df_filter_sum = pd.DataFrame()

        # 라디오 버튼 필터
        if self.rbt0.isChecked():
            # 아무조건 없을때, 전체 가져오기
            df_info_r = df_spr_info["SPR"].reset_index(drop=True)

        if self.rbt1.isChecked():
            info_con = df_spr_info["파트너"] == self.ed1_1.text()
            df_info_r = df_spr_info[info_con]["SPR"].reset_index(drop=True)

        if self.rbt2.isChecked():
            info_con = df_spr_info["장비사"] == self.ed1_1.text()
            df_info_r = df_spr_info[info_con]["SPR"].reset_index(drop=True)

        if self.rbt3.isChecked():
            info_con = df_spr_info["SPR"] == self.ed1_1.text()
            df_info_r = df_spr_info[info_con]["SPR"].reset_index(drop=True)

        if self.rbt4.isChecked():
            info_con = df_spr_info["LGES"] == self.ed1_1.text()
            df_info_r = df_spr_info[info_con]["SPR"].reset_index(drop=True)

        if self.rbt5.isChecked():
            info_con1 = df_spr_info["SPR_Req"] > self.ed1_2.text()
            info_con2 = df_spr_info["SPR_Req"] < self.ed1_3.text()
            info_con = info_con1 & info_con2
            df_info_r = df_spr_info[info_con]["SPR"].reset_index(drop=True)

        # 필터입력에 따른 필터
        if self.ed2_1.text():
            info_con0 = df_spr_info["파트너"] == self.ed2_1.text()
            df_info_r = df_spr_info.loc[info_con0]["SPR"].reset_index(drop=True)

        if self.ed2_2.text():
            info_con1 = df_spr_info["장비사"] == self.ed2_2.text()
            df_info_r = (
                df_spr_info.loc[info_con0].loc[info_con1]["SPR"].reset_index(drop=True)
            )

        if self.ed2_3.text():
            info_con2 = df_spr_info["LGES"] == self.ed2_3.text()
            df_info_r = (
                df_spr_info.loc[info_con0]
                .loc[info_con1]
                .loc[info_con2]["SPR"]
                .reset_index(drop=True)
            )

        if self.ed2_4.text():
            info_con3 = df_spr_info["공정"] == self.ed2_4.text()
            df_info_r = (
                df_spr_info.loc[info_con0]
                .loc[info_con1]
                .loc[info_con2]
                .loc[info_con3]["SPR"]
                .reset_index(drop=True)
            )

        for spr in df_info_r:
            sum_con = df_spr_sum["SPR"] == spr
            if not df_spr_sum[sum_con]["SPR"].empty:
                df_filter_sum = pd.concat([df_filter_sum, df_spr_sum[sum_con]], axis=0)

        # 필터는 SPR INFO 머지하기전 데이터프레임 임.
        df_filter_sum = df_filter_sum.reset_index(drop=True)
        df_merge_sum = df_filter_sum.merge(df_spr_info, how="left").fillna("")

        if debug or D_load_spr_sum_excel:
            SUCCESS = f"load_spr_sum_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE df_info_r: {df_info_r} ")
            print(f"CODE df_filter_sum: {df_filter_sum} ")
            print(f"CODE df_merge_sum: {df_merge_sum} ")

        return df_merge_sum

    except:
        if debug or D_load_spr_sum_excel:
            ERROR = f"load_spr_sum_excel : SPR_PRG_INFO 엑셀파일이없습니다. 새로 생성해주세요."
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : D_load_llp_sum_spr_excel
#
# comment : 필터적용안된 SPR 데이터를 불러옴
###############################################################################################
def load_llp_sum_spr_excel(self):
    drv_path = "D:"
    drv_path = r"Y:\TEST"
    path_spr_info = (
        drv_path
        + r"\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\001. SPR INFO\00. SPR FILTER INFO/"
    )
    file_spr_info = "SPR_PRG_INFO.xlsx"
    src_spr_info = path_spr_info + file_spr_info

    path_spr_sum = drv_path + r"\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\003. SPR SUM/"
    file_spr_sum = "SUM_SPR_DATA.xlsx"
    src_spr_sum = path_spr_sum + file_spr_sum

    try:
        df_spr_info = pd.read_excel(src_spr_info, sheet_name=0)
        df_spr_sum = pd.read_excel(src_spr_sum, sheet_name=0)
        df_filter_sum = pd.DataFrame()

        # 라디오 버튼 필터
        if self.rbt0.isChecked():
            # 아무조건 없을때, 전체 가져오기
            cond = df_spr_sum["MLFB"] == self.ed3_1.text()
            df_llp = df_spr_sum[cond].fillna("").reset_index(drop=True)
            df_merge_sum = df_llp.merge(df_spr_info, how="left").fillna("")

        if self.rbt1.isChecked():
            info_con = df_spr_info["파트너"] == self.ed1_1.text()
            df_info_r = df_spr_info[info_con].reset_index(drop=True)
            cond = df_spr_sum["MLFB"] == self.ed3_1.text()
            df_llp = df_spr_sum[cond].fillna("").reset_index(drop=True)
            df_merge_sum = df_llp.merge(df_info_r, how="left").dropna()

        # cond = df_spr_sum['MLFB'] == self.ed3_1.text()

        # 필터는 SPR INFO 머지하기전 데이터프레임 임.

        if debug or D_load_llp_sum_spr_excel:
            SUCCESS = f"load_llp_sum_spr_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE df_llp: {df_llp} ")
            print(f"CODE df_merge_sum: {df_merge_sum} ")

        return df_llp

    except:
        if debug or D_load_llp_sum_spr_excel:
            ERROR = f"load_llp_sum_spr_excel : 에러발생"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_spr_info_excel
#
# comment :
###############################################################################################
def load_spr_info_excel():
    # path = r"D:\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\001. SPR INFO/"
    path = r"Y:\TEST\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\001. SPR INFO/"

    file = "SPR_INFO.xlsx"
    src = path + file
    try:
        df1 = pd.read_excel(src, sheet_name="INFO_s")

        if debug or D_load_spr_info_excel:
            SUCCESS = f"load_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df1} ")

        return df1

    except:
        if debug or D_load_spr_info_excel:
            ERROR = f"load_time_excel : ERROR 없는 시트에 접근 하였음"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : make_df_spr_info
#
# comment :
###############################################################################################
def make_df_spr_info(self, org_df, filter=0):
    SUM_SPR_FST_COL = 0
    SUM_SPR_END_COL = 100
    SUM_SPR_FST_ROW = 1
    SUM_SPR_END_ROW = 9
    v_arr_sum_jobs_col_addr = []
    v_str_cur_date = org_df.columns[1]

    try:
        for i in range(SUM_SPR_FST_ROW, SUM_SPR_END_ROW + 1):
            v_arr_sum_jobs_col_addr.append(i)

        df_plan_sum = org_df.iloc[
            SUM_SPR_FST_COL:SUM_SPR_END_COL, v_arr_sum_jobs_col_addr
        ].dropna()
        df_plan_sum["SPR_Req"] = df_plan_sum["SPR_Req"].apply(
            lambda x: pd.Timestamp(x).strftime("%Y-%m-%d")
        )

        # df_plan_sum.insert(0,"날짜", v_str_cur_date)

        if filter == Filter.PATHNER:
            cond = df_plan_sum["파트너"] == self.ed1_1.text()
            df_filter = df_plan_sum[cond]
            df_plan_sum = df_filter

        elif filter == Filter.OEM:
            cond = df_plan_sum["장비사"] == self.ed1_1.text()
            df_filter = df_plan_sum[cond]
            df_plan_sum = df_filter

        elif filter == Filter.SPR_NO:
            cond = df_plan_sum["SPR"] == self.ed1_1.text()
            df_filter = df_plan_sum[cond]
            df_plan_sum = df_filter

        elif filter == Filter.PROJECT:
            cond = df_plan_sum["LGES"] == self.ed1_1.text()
            df_filter = df_plan_sum[cond]
            df_plan_sum = df_filter

        elif filter == Filter.DATE:
            cond1 = df_plan_sum["SPR_Req"] > self.ed1_2.text()
            cond2 = df_plan_sum["SPR_Req"] < self.ed1_3.text()
            cond_sum = cond1 & cond2
            df_filter = df_plan_sum[cond_sum]
            df_plan_sum = df_filter

        # 필터입력에 따른 필터
        if self.ed2_1.text():
            cond0 = df_plan_sum["파트너"] == self.ed2_1.text()
            df_plan_sum = df_plan_sum.loc[cond0]

        if self.ed2_2.text():
            cond1 = df_plan_sum["장비사"] == self.ed2_2.text()
            df_plan_sum = df_plan_sum.loc[cond0].loc[cond1]

        if self.ed2_3.text():
            cond2 = df_plan_sum["LGES"] == self.ed2_3.text()
            df_plan_sum = df_plan_sum.loc[cond0].loc[cond1].loc[cond2]

        if self.ed2_4.text():
            cond3 = df_plan_sum["공정"] == self.ed2_4.text()
            df_plan_sum = df_plan_sum.loc[cond0].loc[cond1].loc[cond2].loc[cond3]

        if debug or D_make_df_spr_info:
            SUCCESS = f"make_df_spr_info : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df_plan_sum} ")

        return df_plan_sum

    except:
        if debug or D_make_df_spr_info:
            ERROR = f"make_df_spr_info : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_spr_in_folds
#
# comment : 폴더내 엑셀파일을 읽어서 원하는 형태의 DATAFRAME 을 만들어, 배열로 돌려준다.
###############################################################################################
def load_spr_in_folds():
    path = r"D:\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\002. SPR DB\SPR FROM DIOS RAW/"
    path = (
        r"Y:\TEST\97. 업무공유파일\004. 영업\002. SPR 특가요청 DIOS\002. SPR DB\SPR FROM DIOS RAW/"
    )
    frame = []
    col_name = [
        "SPR",
        "MLFB",
        "Option",
        "PG",
        "BU",
        "LLP",
        "Factor",
        "Price",
        "Qty",
        "Total",
        "Dis",
        "Description",
    ]

    try:
        for (dirpath, dirnames, filenames) in walk(path):
            for file in filenames:
                if not file.startswith("~") and file.endswith(".xlsx"):
                    src = path + file
                    file_name = file[:-9]
                    df = pd.read_excel(src, sheet_name=0)
                    df = df.iloc[:, [3, 7, 8, 15, 9, 11, 5]]
                    df.insert(0, "SPR", file_name)
                    df.insert(2, "Option", "Option")
                    df.insert(4, "BU", "BU")
                    df.insert(6, "Factor", "Factor")
                    df.insert(9, "Total", "TOTAL")
                    df.columns = col_name
                    df["Option"] = df["MLFB"].apply(bu_md.get_mlfb)
                    df["BU"] = df["PG"].apply(bu_md.get_bu)
                    df["Factor"] = (
                        df["Dis"].apply(bu_md.get_factor).apply(lambda x: round(x, 3))
                    )
                    df["Total"] = df["Price"] * df["Qty"]
                    df = df.reset_index(drop=True)

                    frame.append(df)

        if debug or D_load_spr_in_folds:
            SUCCESS = f"load_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {frame} ")

        return frame

    except:
        if debug or D_load_spr_in_folds:
            ERROR = f"load_time_excel : ERROR 없는 시트에 접근 하였음"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : D_make_df_from_arr
#
#
# comment : df 배열을 concat 해서 돌려주는 함수, update 가 아닌 그대로 저장
###############################################################################################
def make_df_from_arr(arr_df_frame):
    v_arr_jobs = arr_df_frame
    # v_df_jobs =[]

    try:
        v_df_jobs = v_arr_jobs[0]

        for i in range(0, len(v_arr_jobs)):
            v_df_jobs = pd.concat([v_df_jobs, v_arr_jobs[i]], axis=0)

        if debug or D_make_df_from_arr:
            SUCCESS = f"make_df_from_arr : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_df_jobs} ")

        return v_df_jobs

    except:
        if debug or D_make_df_from_arr:
            ERROR = f"make_df_from_arr : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE = load_time_excel_from_server()
#
#
# return       : engine
# comment      : 서버 연결, ORM 이용해서 SQL 서버에 연결한다.
###############################################################################################
def load_time_excel_from_server():
    engine = connect_sql_server()
    # v_df_arr = load_time_excel()
    tb_name1 = "plan_sum"
    tb_name2 = "plan_job"
    tb_name3 = "plan_other"
    query1 = f"SELECT * FROM {tb_name1};"
    query2 = f"SELECT * FROM {tb_name2};"
    query3 = f"SELECT * FROM {tb_name3};"

    try:
        df_plan_sum = pd.read_sql(query1, con=engine)
        df_plan_job = pd.read_sql(query2, con=engine)
        df_plan_other = pd.read_sql(query3, con=engine)

        if debug or D_load_time_excel_from_server:
            SUCCESS = f"load_time_excel_from_server : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df_plan_sum} ")
            print(f"CODE : {df_plan_job} ")
            print(f"CODE : {df_plan_other} ")

        return df_plan_sum, df_plan_job, df_plan_other

    except:
        if debug or D_load_time_excel_from_server:
            ERROR = f"load_time_excel_from_server : 서버 연결에 실패했습니다."
            print(f"CODE : {ERROR} ")


# FUNCTION CODE = connect_sql_server()
#
#
# return       : engine
# comment      : 서버 연결, ORM 이용해서 SQL 서버에 연결한다.
###############################################################################################
def connect_sql_server():
    USER = os.getenv("USER")
    PASSWORD = os.getenv("PASSWORD")
    HOST = os.getenv("HOST")
    PORT = os.getenv("PORT")
    DATABASE = os.getenv("DATABASE")

    try:
        connection_string = "mysql+pymysql://%s:%s@%s:%s/%s" % (
            USER,
            PASSWORD,
            HOST,
            PORT,
            DATABASE,
        )
        _engine = create_engine(connection_string)

        if debug or D_connect_sql_server:
            SUCCESS = f"connect_sql_server : SUCCESS"
            print(f"CODE : {SUCCESS} ")

        return _engine

    except:
        if debug or D_connect_sql_server:
            ERROR = f"connect_sql_server : 서버 연결에 실패했습니다."
            print(f"CODE : {ERROR} ")


# FUNCTION CODE = save_time_excel_to_server()
#
#
# return       :
# comment      :
###############################################################################################
def save_time_excel_to_server():
    engine = connect_sql_server()
    v_df_arr = load_time_excel()
    tb_name1 = "plan_sum"
    tb_name2 = "plan_job"
    tb_name3 = "plan_other"
    # query = f'SELECT * FROM {tb_name1};'
    try:

        if debug or D_save_time_excel_to_server:
            SUCCESS = f"save_time_excel_to_server : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_df_arr} ")

        v_df_arr[0].to_sql(tb_name1, con=engine, if_exists="replace", index=False)
        v_df_arr[1].to_sql(tb_name2, con=engine, if_exists="replace", index=False)
        v_df_arr[2].to_sql(tb_name3, con=engine, if_exists="replace", index=False)

    except:
        if debug or D_save_time_excel_to_server:
            ERROR = f"save_time_excel_to_server : ERROR "
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_time_excel
# path = 파일이 있는 경로명
# file = 파일명.xlsx
# src  = path + file
#
# comment : 시간분석테이블에 내용을 읽어온다.
###############################################################################################
def load_time_excel():
    path = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/"
    file = "03. 시간분석테이블.xlsx"
    src = path + file
    try:
        df1 = pd.read_excel(src, sheet_name=0)
        df2 = pd.read_excel(src, sheet_name=1)
        df3 = pd.read_excel(src, sheet_name=2)

        if debug or D_load_time_excel:
            SUCCESS = f"load_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df1} ")
            print(f"CODE : {df2} ")
            print(f"CODE : {df3} ")

        return df1, df2, df3

    except:
        if debug or D_load_time_excel:
            ERROR = f"load_time_excel : ERROR 없는 시트에 접근 하였음"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_cur_plan_excel
# path = 파일이 있는 경로명
# file = 파일명.xlsx
# src  = path + file
# cnt_sht_num : 현재날짜 부터 몇개의 시트를 읽어올것인지 파라미터
# comment : 일일계획표에 지정된 시트를 모두 읽어와서 배열로 리턴한다.
###############################################################################################
def load_cur_plan_excel(cmd, sht_num_cnt=1, date="220701"):
    path = r"D:\97. 업무공유파일\000. 계획/"
    file = "01. 2022 일일계획표.xlsx"
    src = path + file
    frame = []
    try:
        v_str_command = cmd
        v_int_cnt_sht_num = int(sht_num_cnt)
        v_str_date = date
        v_int_start_sht_num = 4
        v_int_tar_sht_num = v_int_start_sht_num + v_int_cnt_sht_num

        if v_str_command == "load_shts_by_number":
            for i in range(v_int_start_sht_num, v_int_tar_sht_num):
                df = pd.read_excel(src, sheet_name=i)
                frame.append(df)
        elif v_str_command == "load_sht_by_date":
            df = pd.read_excel(src, sheet_name=v_str_date)
            frame.append(df)

        if debug or D_load_cur_plan_excel:
            SUCCESS = f"load_cur_plan_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            # 디버그시 현재 날짜 출력
            for i in range(0, len(frame)):
                print(f"\tREAD SHT DATE : {frame[i].columns[1]}")

        return frame

    except:
        if debug or D_load_cur_plan_excel:
            ERROR = f"load_cur_plan_excel : ERROR 없는 시트에 접근 하였음"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : make_df_plan_sum
# path
# file        :
# comment :
###############################################################################################
def make_df_plan_sum(org_df):
    SUM_JOBS_FST_COL = 11
    SUM_JOBS_END_COL = 18
    SUM_JOBS_FST_ROW = 21
    SUM_JOBS_END_ROW = 22
    v_arr_sum_jobs_col_addr = []
    v_str_cur_date = org_df.columns[1]

    try:
        for i in range(SUM_JOBS_FST_COL, SUM_JOBS_END_COL + 1):
            v_arr_sum_jobs_col_addr.append(i)

        df_plan_sum = org_df.iloc[
            SUM_JOBS_FST_ROW:SUM_JOBS_END_ROW, v_arr_sum_jobs_col_addr
        ].dropna()
        df_plan_sum.insert(0, "날짜", v_str_cur_date)

        if debug or D_make_df_plan_sum:
            SUCCESS = f"make_df_plan_sum : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df_plan_sum} ")

        return df_plan_sum

    except:
        if debug or D_make_df_plan_sum:
            ERROR = f"make_df_plan_sum : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


def make_df_plan_jobs(org_df):
    v_arr_jobs_col_addr = []
    v_str_cur_date = org_df.columns[1]
    JOBS_FST_COL = 11
    JOBS_END_COL = 24
    JOBS_STEP = 2
    JOBS_FST_ROW = 24
    JOBS_QTY = 7
    v_arr_jobs = []

    try:
        # Get jobs columns array
        for i in range(JOBS_FST_COL, JOBS_END_COL, JOBS_STEP):
            j = i + 1
            v_arr_jobs_col_addr.append([i, j])

        # Get jobs dataframe array
        for i in range(0, JOBS_QTY):
            df_jobs = org_df.iloc[JOBS_FST_ROW:, v_arr_jobs_col_addr[i]].dropna()
            df_jobs = df_jobs.reset_index(drop=True).T.reset_index(drop=True)

            v_arr_jobs.append(df_jobs)

        # Set jobs column name with first row value
        for i in range(0, len(v_arr_jobs)):
            v_arr_jobs[i].columns = v_arr_jobs[i].iloc[0]
            v_arr_jobs[i] = v_arr_jobs[i].drop([v_arr_jobs[i].index[0]])

        df_job_others = pd.concat(
            [
                v_arr_jobs[1],
                v_arr_jobs[2],
                v_arr_jobs[3],
                v_arr_jobs[4],
                v_arr_jobs[5],
                v_arr_jobs[6],
            ],
            axis=1,
        )

        # insert 날짜
        v_arr_jobs[0].insert(0, "날짜", v_str_cur_date)
        df_job_others.insert(0, "날짜", v_str_cur_date)

        if debug or D_make_df_plan_jobs:
            SUCCESS = f"make_df_plan_jobs : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_arr_jobs[0]} ")
            print(f"CODE : {df_job_others} ")

        return v_arr_jobs[0], df_job_others

    except:
        if debug or D_make_df_plan_jobs:
            ERROR = f"make_df_plan_jobs : ERROR 데이터 프레임 만들기 실패"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : D_update_df_from_arr
#
#
# comment : df 배열을 concat 해서 돌려주는 함수
###############################################################################################
def update_df_from_arr(arr_df_frame):
    v_arr_jobs = arr_df_frame
    v_df_jobs = [[], [], []]
    v_df_time_jobs = load_time_excel()

    try:
        for i in range(0, len(v_arr_jobs)):
            v_df_jobs[i] = v_df_time_jobs[i]
            for j in range(0, len(v_arr_jobs[i])):
                v_df_jobs[i] = pd.concat([v_df_jobs[i], v_arr_jobs[i][j]], axis=0)
                v_df_jobs[i] = v_df_jobs[i].drop_duplicates("날짜").reset_index(drop=True)
                v_df_jobs[i] = v_df_jobs[i].sort_values("날짜", ascending=False)

        if debug or D_update_df_from_arr:
            SUCCESS = f"update_df_from_arr : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_df_jobs[0]} ")

        return v_df_jobs

    except:
        if debug or D_update_df_from_arr:
            ERROR = f"update_df_from_arr : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : save_time_excel
# path
# file        :
# comment :
###############################################################################################


def save_time_excel(df1, df2, df3):
    src = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/"
    file = "03. 시간분석테이블.xlsx"
    file_b = "03. 시간분석테이블-backup.xlsx"
    src_file = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/03. 시간분석테이블.xlsx"
    sht = []
    sht = read_sheets(src, file)

    try:
        if debug or D_save_time_excel:
            SUCCESS = f"save_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {sht} ")

        with pd.ExcelWriter(src_file) as writer:
            df1.to_excel(writer, sheet_name=sht[0], index=False)
            df2.to_excel(writer, sheet_name=sht[1], index=False)
            df3.to_excel(writer, sheet_name=sht[2], index=False)

        with pd.ExcelWriter(src + file_b) as writer:
            df1.to_excel(writer, sheet_name=sht[0], index=False)
            df2.to_excel(writer, sheet_name=sht[1], index=False)
            df3.to_excel(writer, sheet_name=sht[2], index=False)

    except:
        if debug or D_save_time_excel:
            ERROR = f"save_time_excel : ERROR 타임시트 엑셀에 저장실패"
            print(f"CODE : {ERROR} ")
